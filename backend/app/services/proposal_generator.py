"""Generates downloadable Estimate Sheet (xlsx) and Proposal summary (pdf)."""
import io
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill
from reportlab.lib.pagesizes import A4
from reportlab.lib.units import cm
from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer
from reportlab.lib.styles import getSampleStyleSheet
from reportlab.lib import colors


def build_estimate_sheet_xlsx(cost_sheet: dict) -> bytes:
    wb = Workbook()
    ws = wb.active
    ws.title = "Estimate Sheet"

    headers = ["Item Code", "Description", "Qty", "Unit", "Unit Cost", "Unit Price", "Line Cost", "Line Price", "Margin %"]
    ws.append(headers)
    for cell in ws[1]:
        cell.font = Font(bold=True, color="FFFFFF")
        cell.fill = PatternFill("solid", fgColor="1F4E78")

    for item in cost_sheet["line_items"]:
        ws.append([
            item.item_code, item.description, item.quantity, item.unit,
            item.unit_cost, item.unit_price, item.line_cost, item.line_price, item.margin_percent
        ])

    ws.append([])
    ws.append(["", "", "", "", "", "Subtotal (Cost)", cost_sheet["subtotal_cost"]])
    ws.append(["", "", "", "", "", "Subtotal (Price)", cost_sheet["subtotal_price"]])
    ws.append(["", "", "", "", "", f"Contingency ({cost_sheet['contingency_percent']}%)", cost_sheet["contingency_amount"]])
    ws.append(["", "", "", "", "", "Grand Total", cost_sheet["grand_total_price"]])

    for col in "ABCDEFGHI":
        ws.column_dimensions[col].width = 18

    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()


def build_proposal_pdf(project_name: str, customer_name: str, cost_sheet: dict) -> bytes:
    buf = io.BytesIO()
    doc = SimpleDocTemplate(buf, pagesize=A4, topMargin=2*cm, bottomMargin=2*cm)
    styles = getSampleStyleSheet()
    elements = [
        Paragraph(f"Commercial Proposal — {project_name}", styles["Title"]),
        Paragraph(f"Prepared for: {customer_name or 'Customer'}", styles["Normal"]),
        Spacer(1, 0.5*cm),
    ]

    data = [["Description", "Qty", "Unit", "Unit Price", "Line Price"]]
    for item in cost_sheet["line_items"]:
        data.append([item.description, str(item.quantity), item.unit, f"{item.unit_price:.2f}", f"{item.line_price:.2f}"])
    data.append(["", "", "", "Subtotal", f"{cost_sheet['subtotal_price']:.2f}"])
    data.append(["", "", "", f"Contingency ({cost_sheet['contingency_percent']}%)", f"{cost_sheet['contingency_amount']:.2f}"])
    data.append(["", "", "", "Grand Total", f"{cost_sheet['grand_total_price']:.2f} {cost_sheet['currency']}"])

    table = Table(data, colWidths=[7*cm, 2*cm, 2*cm, 3*cm, 3*cm])
    table.setStyle(TableStyle([
        ("BACKGROUND", (0, 0), (-1, 0), colors.HexColor("#1F4E78")),
        ("TEXTCOLOR", (0, 0), (-1, 0), colors.white),
        ("FONTSIZE", (0, 0), (-1, -1), 8),
        ("GRID", (0, 0), (-1, -4), 0.5, colors.grey),
        ("FONTNAME", (3, -1), (-1, -1), "Helvetica-Bold"),
    ]))
    elements.append(table)
    doc.build(elements)
    return buf.getvalue()
